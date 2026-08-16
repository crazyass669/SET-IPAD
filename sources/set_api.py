# -*- coding: utf-8 -*-
"""
sources/set_api.py — client สำหรับ SET internal API (www.set.or.th)

ใช้เป็น primary ของ fundamentals (mkt_cap/PE/PBV/DivYield) แทน Yahoo ที่โดน
rate limit บ่อย — และเป็นโครงสำหรับ backup ราคาในอนาคต (chart-quotation)

หมายเหตุ: เป็น internal API ของเว็บ SET (ไม่มีสัญญา) — ทุกฟังก์ชันมี
validation gate: ถ้าผลได้ต่ำกว่าเกณฑ์ให้ raise เพื่อให้ caller fallback
ไปแหล่งอื่นแทนที่จะรับข้อมูลไม่ครบเข้า pipeline เงียบๆ
"""
import json
import time
import urllib.parse
import urllib.request as ur
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd

from core.net import ssl_context

BASE = "https://www.set.or.th"
_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/125.0"


def _bootstrap_headers(referer_path="/th/market/product/stock/quote/PTT/price"):
    """เปิดหน้าเว็บหนึ่งครั้งเพื่อรับ cookie — internal API ตอบ 403 ถ้าไม่มี"""
    ctx = ssl_context()
    req = ur.Request(BASE + referer_path, headers={"User-Agent": _UA})
    with ur.urlopen(req, context=ctx, timeout=20) as r:
        cookie = r.getheader("Set-Cookie", "") or ""
    return ctx, {
        "User-Agent": _UA,
        "Accept": "application/json",
        "Referer": BASE + referer_path,
        "Cookie": cookie,
    }


def _get_json(ctx, hdr, path, timeout=12):
    req = ur.Request(BASE + path, headers=hdr)
    with ur.urlopen(req, context=ctx, timeout=timeout) as r:
        return json.loads(r.read().decode("utf-8", "ignore"))


def fetch_fundamentals(tickers, callback=None, workers=8, min_ratio=0.5):
    """ดึง mkt_cap/PE/PBV/DivYield จาก /api/set/stock/<sym>/highlight-data

    tickers: list ของ "PTT.BK" (รูปแบบเดียวกับ fetch_market_caps_parallel)
    คืน {ticker: {"mkt_cap": int|None, "pe": float|None,
                  "pbv": float|None, "div_yield": float|None}}
    raise ValueError ถ้าดึงสำเร็จน้อยกว่า min_ratio (ให้ caller fallback Yahoo)
    """
    ctx, hdr = _bootstrap_headers()
    results = {}

    def _one(tick):
        sym = tick[:-3] if tick.endswith(".BK") else tick
        path = f"/api/set/stock/{urllib.parse.quote(sym)}/highlight-data?lang=th"
        for attempt in range(2):
            try:
                d = _get_json(ctx, hdr, path)
                mc  = d.get("marketCap")
                pe  = d.get("peRatio")
                pbv = d.get("pbRatio")
                dy  = d.get("dividendYield")
                return tick, {
                    "mkt_cap":   int(mc)            if mc  is not None else None,
                    "pe":        round(float(pe), 2)  if pe  is not None else None,
                    "pbv":       round(float(pbv), 2) if pbv is not None else None,
                    "div_yield": round(float(dy), 2)  if dy  is not None else None,
                }
            except Exception:
                if attempt == 0:
                    time.sleep(0.5)
        return tick, None

    total = len(tickers)
    done = ok = 0
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = [ex.submit(_one, t) for t in tickers]
        for f in as_completed(futures):
            tick, data = f.result()
            done += 1
            if data is not None:
                results[tick] = data
                ok += 1
            if callback and done % 100 == 0:
                callback(done, total, f"Fundamentals (SET API) {done}/{total}...")

    if total > 0 and ok < total * min_ratio:
        raise ValueError(f"SET API fundamentals ได้แค่ {ok}/{total} "
                         f"(<{int(min_ratio*100)}%) — ควร fallback")
    return results


_SUSPEND_SIGNS = {"SP", "NC", "NP", "H"}


def fetch_quotes_batch(tickers, callback=None, workers=6, batch_size=30, min_ratio=0.5,
                        signs_out=None):
    """ดึงราคาล่าสุด (OHLCV วันเดียว) ของทั้งกระดานเร็ว จาก
    /api/set/stock-info/list-by-symbols — primary ของ Quick Update แทน Yahoo
    (เร็วกว่ามาก + ไม่ lag เหมือน Yahoo ที่ปล่อยแท่งปิดหุ้นไทยช้า — ดู
    services/refresh.py::run_quick_update สำหรับ fast-path logic เต็ม)

    tickers: list ของ "PTT.BK" -> คืน {ticker: {"open","high","low","close",
             "volume","prior"}} เฉพาะตัวที่มีราคาจริง (ข้าม sign SP/NC/NP/H หรือ
             OHLC เป็น null — หุ้นพักเทรด/non-compliant)
    signs_out: dict ว่างที่ส่งเข้ามาให้ฟังก์ชันเติมให้ (ticker -> sign string ดิบ เช่น
             "SP,NC,NP") เฉพาะตัวที่ SET คืนเครื่องหมายมาไม่ว่าง — เดิมอ่าน sign มาเช็ค
             แล้วทิ้งเงียบๆ ไม่มีทางรู้ว่าตัวที่หายไปจากผลลัพธ์เป็นเพราะเครื่องหมายอะไร
             (ดู detect_flat_price_tickers ใน services/refresh.py ที่ใช้ param นี้)
    รับสูงสุด 30 ตัว/request (เกินกว่านั้นตอบ 400)
    raise ValueError ถ้าดึงสำเร็จน้อยกว่า min_ratio (ให้ caller fallback Yahoo)
    """
    ctx, hdr = _bootstrap_headers()
    sym_map = {(t[:-3] if t.endswith(".BK") else t): t for t in tickers}
    chunks = [tickers[i:i + batch_size] for i in range(0, len(tickers), batch_size)]
    results = {}

    def _one(chunk):
        syms = [t[:-3] if t.endswith(".BK") else t for t in chunk]
        q = urllib.parse.quote(",".join(syms))
        for attempt in range(2):
            try:
                return _get_json(ctx, hdr,
                                  f"/api/set/stock-info/list-by-symbols?symbols={q}&lang=th")
            except Exception:
                if attempt == 0:
                    time.sleep(0.5)
        return None

    total = len(tickers)
    ok = 0
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = [ex.submit(_one, c) for c in chunks]
        for f in as_completed(futures):
            rows = f.result()
            if not rows:
                continue
            for r in rows:
                tick = sym_map.get(r.get("symbol"))
                if not tick:
                    continue
                sign_raw = (r.get("sign") or "").strip()
                if sign_raw and signs_out is not None:
                    signs_out[tick] = sign_raw
                sign_codes = {c.strip() for c in sign_raw.split(",") if c.strip()}
                o, h, l, c = r.get("open"), r.get("high"), r.get("low"), r.get("last")
                if sign_codes & _SUSPEND_SIGNS or None in (o, h, l, c):
                    continue
                prior = r.get("prior")
                results[tick] = {
                    "open": float(o), "high": float(h), "low": float(l), "close": float(c),
                    "volume": int(r.get("totalVolume") or 0),
                    "prior": float(prior) if prior is not None else None,
                }
                ok += 1
            if callback:
                callback(ok, total, f"ราคาล่าสุด (SET API) {ok}/{total}...")

    if total > 0 and ok < total * min_ratio:
        raise ValueError(f"SET API quotes ได้แค่ {ok}/{total} "
                         f"(<{int(min_ratio*100)}%) — ควร fallback")
    return results


def fetch_signs_batch(tickers, workers=6, batch_size=30):
    """เช็คเฉพาะเครื่องหมาย (sign) ปัจจุบันของหุ้นกลุ่มเล็กที่ระบุ ไม่สนราคา — ใช้กับ ticker
    ที่ตรวจพบว่าราคาปิดคงที่ผิดปกติหลายวันติด (ดู detect_flat_price_tickers ใน
    services/refresh.py) เพื่อยืนยันว่าเป็นเพราะ SP/NC/NP/H จริงไหม min_ratio=0 เพราะ
    ตัวที่ต้องสงสัยส่วนใหญ่คาดว่าจะถูกกรองออกจาก results อยู่แล้ว (นั่นคือประเด็นที่จะเช็ค)
    ไม่ควร raise แค่เพราะ "ได้ราคาน้อย" — คืน {ticker: sign string} เฉพาะตัวที่มีเครื่องหมาย"""
    signs = {}
    try:
        fetch_quotes_batch(tickers, workers=workers, batch_size=batch_size, min_ratio=0,
                            signs_out=signs)
    except Exception:
        pass
    return signs


def fetch_trading_calendar_tail(ref_symbol="PTT", n=5):
    """คืนวันเทรดล่าสุด n วัน ('YYYY-MM-DD', เรียงเก่า->ใหม่) จาก chart-quotation ของ
    หุ้นอ้างอิงตัวเดียว (blue chip สภาพคล่องสูง แทบไม่มีพักเทรด) — list-by-symbols
    (fetch_quotes_batch) ไม่มี field วันที่ตรงๆ ในตัวเอง (marketDateTime คือเวลาที่
    query ไม่ใช่วันเทรด — ยืนยันแล้วว่าต่างกันได้เวลา query ก่อนตลาดเปิด/วันหยุด)

    เอา n>=2 เพื่อให้ caller เช็คได้ว่า 'วันก่อนวันล่าสุด' ตรงกับวันที่เก็บไว้จริงไหม —
    กัน Quick Update fast path เข้าใจผิดว่ามีช่องว่างแค่ 1 วัน ทั้งที่จริงขาดหลายวัน
    (ดู services/refresh.py::run_quick_update) raise ถ้าดึงไม่ได้"""
    bars = fetch_daily_bars(ref_symbol, period="1M")
    if not bars:
        raise ValueError("ดึงปฏิทินวันเทรดล่าสุดไม่ได้ (chart-quotation ว่างเปล่า)")
    return [b[0] for b in bars[-n:]]


def fetch_daily_bars(symbol, period="1M", ctx=None, hdr=None):
    """สำรองราคา: ราคา+volume รายวันของหุ้นหนึ่งตัวจาก chart-quotation
    คืน list ของ (date_str, close, volume) — สำหรับ emergency backup ของ Yahoo

    ctx/hdr: ส่งมาเองได้เพื่อ reuse cookie เดียวข้ามหลายเรียก (ดู
    fetch_price_history_batch ที่เรียกฟังก์ชันนี้เป็นร้อยครั้งแบบ threaded — ไม่งั้น
    จะ bootstrap คุกกี้ใหม่ทุกครั้งซึ่งช้าเกินจำเป็น) ไม่ส่งมาก็ bootstrap ให้เองเหมือนเดิม"""
    if ctx is None or hdr is None:
        ctx, hdr = _bootstrap_headers()
    sym = symbol[:-3] if symbol.endswith(".BK") else symbol
    d = _get_json(ctx, hdr,
                  f"/api/set/stock/{urllib.parse.quote(sym)}/chart-quotation"
                  f"?period={period}&accumulated=false")
    out = []
    for q in d.get("quotations", []):
        dt = (q.get("localDatetime") or q.get("datetime") or "")[:10]
        px = q.get("price")
        if dt and px is not None:
            out.append((dt, float(px), int(q.get("volume") or 0)))
    return out


_PERIOD_TIERS = (("1M", 28), ("3M", 88), ("6M", 178), ("1Y", 360))


def fetch_price_history_batch(tickers, start_date, callback=None, workers=6):
    """สำรองราคาย้อนหลังหลายวันจาก chart-quotation ทั้งกระดาน — fallback ชั้นสุดท้าย
    ของ Quick Update เมื่อ Yahoo ล่ม/ไม่ตอบหลายวันติด (ปกติ Yahoo เป็นตัวเติม gap
    หลายวัน — ดู services/refresh.py::run_quick_update)

    ต่างจาก fetch_quotes_batch (list-by-symbols) ตรงที่ได้ "ประวัติย้อนหลัง" จริงหลายวัน
    (ไม่ใช่แค่วันล่าสุดวันเดียว) แต่ chart-quotation ไม่มี field OHLC เลย ได้แค่
    close+volume — พอให้ราคา/RS/เทคนิคขยับต่อได้ระหว่างรอ Yahoo ฟื้น พอ Yahoo ตอบปกติ
    Quick Update รอบถัดไปจะเติม OHLC ให้เองผ่าน Yahoo (upsert_bars รองรับ OHLC เป็น
    NULL อยู่แล้ว เหมือนแถวเก่าก่อน migration — ดู core/store.py init_db)

    tickers: list ของ "PTT.BK", start_date: "YYYY-MM-DD"
    คืน {ticker: {"close": pd.Series, "volume": pd.Series}} เฉพาะตัวที่ดึงได้ — ไม่ raise
    ถ้าบางตัวพลาด (เป็น fallback ชั้นสุดท้ายอยู่แล้ว ได้เท่าไหร่เอาเท่านั้น)"""
    days_needed = max((pd.Timestamp.now().normalize() - pd.to_datetime(start_date)).days, 1)
    period = "1Y"
    for p, max_days in _PERIOD_TIERS:
        if days_needed <= max_days:
            period = p
            break

    ctx, hdr = _bootstrap_headers()
    results = {}

    def _one(tick):
        sym = tick[:-3] if tick.endswith(".BK") else tick
        try:
            bars = fetch_daily_bars(sym, period=period, ctx=ctx, hdr=hdr)
        except Exception:
            return tick, None
        keep = [(d, c, v) for d, c, v in bars if d >= start_date]
        if not keep:
            return tick, None
        idx = pd.DatetimeIndex([pd.Timestamp(d) for d, _, _ in keep])
        close  = pd.Series([c for _, c, _ in keep], index=idx, dtype=float)
        volume = pd.Series([v for _, _, v in keep], index=idx, dtype=float)
        return tick, {"close": close, "volume": volume}

    total = len(tickers)
    done = ok = 0
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = [ex.submit(_one, t) for t in tickers]
        for f in as_completed(futures):
            tick, data = f.result()
            done += 1
            if data is not None:
                results[tick] = data
                ok += 1
            if callback and done % 100 == 0:
                callback(done, total, f"สำรองราคา (SET API chart-quotation) {done}/{total}...")

    return results


# ============================================================
# งบกำไรขาดทุนรายไตรมาสละเอียด (สำหรับตาราง P&L รายไตรมาส) — เจอ endpoint พวกนี้
# จากการไล่อ่าน JS bundle ของ settrade.com (ใช้ backend เดียวกับ set.or.th ผ่าน
# /api/set/... ) ไม่มีเอกสารทางการรองรับเหมือน endpoint อื่นในไฟล์นี้ทั้งหมด
# ============================================================

def fetch_financial_data_chart(symbol, ctx=None, hdr=None):
    """สรุปรายไตรมาส 'เดี่ยวๆ' (ไม่สะสม) ย้อนหลัง ~5 ปี จาก company-highlight/financial-data-chart
    (accumulated=false) — ต่างจาก company-highlight ปกติตรงที่ไม่มีปัญหางวด '6M'/'9M' ปนมา
    (บริษัทที่ยื่นงบล่าสุดเป็นงวดสะสม เช่น PTTEP/AOT) เพราะ endpoint นี้ตัดเป็นรายไตรมาสให้เอง
    field มีแค่ระดับสรุป (revenue/totalExpense/GPM%/ebit/netProfit) ไม่มี COGS/SG&A แยก
    คืน list ของ raw dict เรียงเก่า->ใหม่ เฉพาะ quarter ที่เป็น Q1-Q4 (กันกรณี edge ที่หลุดมา)"""
    sym = symbol[:-3] if symbol.endswith(".BK") else symbol
    if ctx is None or hdr is None:
        ctx, hdr = _bootstrap_headers()
    d = _get_json(ctx, hdr,
                  f"/api/set/stock/{urllib.parse.quote(sym)}/company-highlight/financial-data-chart"
                  f"?accumulated=false&lang=th", timeout=15)
    return [r for r in (d or []) if r.get("quarter") in ("Q1", "Q2", "Q3", "Q4")]


def fetch_financial_statement_periods(symbol, ctx=None, hdr=None):
    """รายชื่องวดที่มีงบละเอียดให้ดึงได้ (เช่น ['Q1_2026','YE_2025','9M_2025','6M_2025']) —
    มีแค่ปีปัจจุบัน+ปีก่อนหน้าเท่านั้น ไม่ย้อนลึกหลายปี (เช็คกับ PTT/IRPC/CH/LPH/SIAM แล้ว)"""
    sym = symbol[:-3] if symbol.endswith(".BK") else symbol
    if ctx is None or hdr is None:
        ctx, hdr = _bootstrap_headers()
    d = _get_json(ctx, hdr,
                  f"/api/set/stock/{urllib.parse.quote(sym)}/financialstatement/periods", timeout=15)
    return (d or {}).get("periods") or []


def fetch_financial_statement(symbol, account_type="income_statement", period=None, ctx=None, hdr=None):
    """งบละเอียดทีละบรรทัดบัญชี (accountCode/accountName/amount) ตรงกับที่บริษัทยื่นจริง —
    account_type: income_statement | balance_sheet | cash_flow
    period: None = งวดล่าสุด, หรือค่าจาก fetch_financial_statement_statement_periods()
    (เช่น 'YE_2025') เพื่อดึงงวดอื่นที่ไม่ใช่ล่าสุด — คืน raw dict {symbol,year,quarter,
    endDate,accounts:[...]} unit เป็น 'พันบาท' (divider=1000 ทุกแถว)"""
    sym = symbol[:-3] if symbol.endswith(".BK") else symbol
    if ctx is None or hdr is None:
        ctx, hdr = _bootstrap_headers()
    params = {"accountType": account_type, "lang": "th"}
    if period:
        params["period"] = period
    q = urllib.parse.urlencode(params)
    return _get_json(ctx, hdr,
                      f"/api/set/stock/{urllib.parse.quote(sym)}/financialstatement?{q}", timeout=15)


def fetch_trading_halts(ctx=None, hdr=None):
    """รายชื่อหลักทรัพย์ที่หยุดพักการซื้อขายชั่วคราว (SP/H/P) ปัจจุบัน — ขุด endpoint จาก
    JS bundle (Nuxt) ของหน้า
    https://www.set.or.th/th/market/news-and-alert/trading-halt-suspension-pause
    (หน้านั้น render ฝั่ง client ล้วน ดึง HTML ตรงๆ ไม่เห็นข้อมูล เหมือน delisted-list)

    ไม่ส่ง query param ใดๆ = ดึงทุก securityType/sign (ตรงกับปุ่ม "ทั้งหมด" บนหน้าเว็บ)
    securityType 'S' คือหุ้นสามัญ (ใช้ symbol ตรงกับที่ใช้ทั้ง dashboard ไม่มี suffix) —
    'F'/'W' คือกระดานต่างชาติ/วอแรนท์ของบริษัทเดียวกัน ปกติหยุดพักพร้อมกับ 'S' เสมอ

    คืน {"as_of": "...ISO datetime...", "items": [{"symbol","sign","name_th","name_en",
    "security_type","market"}, ...]}"""
    if ctx is None or hdr is None:
        ctx, hdr = _bootstrap_headers(
            referer_path="/th/market/news-and-alert/trading-halt-suspension-pause")
    d = _get_json(ctx, hdr, "/api/set/market/trading-halts/list", timeout=15)
    items = [{
        "symbol": r.get("symbol"),
        "sign": r.get("sign"),
        "name_th": r.get("nameTH"),
        "name_en": r.get("nameEN"),
        "security_type": r.get("securityType"),
        "market": r.get("market"),
    } for r in (d or {}).get("stockList") or []]
    return {"as_of": (d or {}).get("datetime") or "", "items": items}


def fetch_delisted_companies(lang="th", ctx=None, hdr=None):
    """รายชื่อหุ้นเพิกถอนที่ SET ยืนยันแล้วทั้งหมด (official, ไม่ใช่ heuristic เดา
    จากราคานิ่ง) — ขุด endpoint จาก JS bundle ของหน้า
    https://www.set.or.th/th/market/information/securities-list/delisted-list
    (หน้านั้น render ฝั่ง client ล้วน ดึง HTML ตรงๆ ไม่เห็นข้อมูล)

    คืน {"as_of_date": "YYYY-MM-DD", "companies": [{"symbol","name","market",
    "delist_date","delist_reason"}, ...]} ย้อนหลังถึงปี 1976 (~340 ตัว)"""
    if ctx is None or hdr is None:
        ctx, hdr = _bootstrap_headers()
    d = _get_json(ctx, hdr, f"/api/set/company/delisted-company?lang={lang}", timeout=20)
    comps = (d or {}).get("delistedCompanies") or []
    return {
        "as_of_date": (d.get("asOfDate") or "")[:10],
        "companies": [{
            "symbol": c.get("symbol"),
            "name": c.get("name"),
            "market": c.get("market"),
            "delist_date": (c.get("delistDate") or "")[:10],
            "delist_reason": c.get("delistReason") or "",
        } for c in comps],
    }


_TH_MONTHS = {
    "ม.ค.": 1, "ก.พ.": 2, "มี.ค.": 3, "เม.ย.": 4, "พ.ค.": 5, "มิ.ย.": 6,
    "ก.ค.": 7, "ส.ค.": 8, "ก.ย.": 9, "ต.ค.": 10, "พ.ย.": 11, "ธ.ค.": 12,
}


def _parse_thai_date(s):
    """'13 ธ.ค. 2567' (พ.ศ.) -> '2024-12-13' (ค.ศ.) — คืน '' ถ้า parse ไม่ได้/เป็น '-'"""
    if not s or not isinstance(s, str):
        return ""
    parts = s.strip().split()
    if len(parts) != 3:
        return ""
    d, mo, y = parts
    month = _TH_MONTHS.get(mo)
    if not month:
        return ""
    try:
        return f"{int(y) - 543:04d}-{month:02d}-{int(d):02d}"
    except ValueError:
        return ""


def parse_delisted_companies_xlsx(path):
    """อ่านไฟล์ 'delisted-securities-th.xlsx' ที่ดาวน์โหลดมาด้วยมือจากหน้า
    https://www.set.or.th/th/market/information/securities-list/delisted-list
    (ปุ่มดาวน์โหลดบนหน้านั้น) — ทางเลือกแทน fetch_delisted_companies() ที่ต้องผ่าน
    bootstrap cookie/WAF คอลัมน์: หลักทรัพย์, ชื่อบริษัท, ตลาด, กลุ่มอุตสาหกรรม,
    หมวดธุรกิจ, วันที่ซื้อขายวันแรก, วันที่เพิกถอน, สาเหตุ (แถว 1-2 = หัวเรื่อง/หัวตาราง)

    คืน shape เดียวกับ fetch_delisted_companies(): {"as_of_date", "companies":[...]}"""
    import re

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    as_of_date = ""
    if rows:
        header_text = " ".join(str(c) for c in rows[0] if c)
        m = re.search(r"(\d{1,2})\s+(\S+\.)\s+(\d{4})", header_text)
        if m:
            as_of_date = _parse_thai_date(f"{m.group(1)} {m.group(2)} {m.group(3)}")

    companies = []
    for row in rows[2:]:
        if not row or not row[0]:
            continue
        companies.append({
            "symbol": str(row[0]).strip(),
            "name": str(row[1]).strip() if row[1] else "",
            "market": str(row[2]).strip() if row[2] else "",
            "delist_date": _parse_thai_date(row[6]),
            "delist_reason": str(row[7]).strip() if row[7] else "",
        })
    return {"as_of_date": as_of_date, "companies": companies}
